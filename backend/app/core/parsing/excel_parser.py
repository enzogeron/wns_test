from dataclasses import dataclass
from typing import Any, List, Optional

from app.core.parsing.base import Parser
from app.core.parsing.types import DocumentMeta, SheetTable
from openpyxl import load_workbook


@dataclass(frozen=True)
class ExcelDocument:
    meta: DocumentMeta
    tables: List[SheetTable]

class ExcelTableParser(Parser[ExcelDocument]):
    def parse(self, path: str, source_name: Optional[str] = None) -> ExcelDocument:
        meta = self._meta(path, source_name, "ExcelTableParser")

        wb = load_workbook(path, data_only=True)
        tables: List[SheetTable] = []

        for ws in wb.worksheets:
            rows: List[List[Any]] = []
            max_cols = 0

            for r in ws.iter_rows(values_only=True):
                row = list(r)
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                max_cols = max(max_cols, len(row))
                rows.append(row)

            tables.append(
                SheetTable(
                    sheet_name=ws.title,
                    rows=rows,
                    meta={"max_cols": max_cols, "rows": len(rows)},
                )
            )

        return ExcelDocument(meta=meta, tables=tables)